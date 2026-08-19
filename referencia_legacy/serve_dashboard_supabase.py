from __future__ import annotations

import io
import json
import os
import re
import shutil
import threading
import time
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from email.parser import BytesParser
from email.policy import default as email_policy
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib import error as urlerror
from urllib import request as urlrequest
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd

from build_biofisic_tabs_dashboard import (
    ANALYSIS_SNAPSHOT_FILE,
    DEFAULT_CSV_DIR,
    HTML_OUT,
    OUT_DIR,
    UNIT_ORDER,
    build_blank_payload,
    build_html,
    build_payload,
    load_processed_tables,
    prepare_table_frames,
    source_upload_label,
)
from supabase_data_source import SupabaseConfig, SupabaseTableLoader, supabase_source_label

try:
    import cgi
except ModuleNotFoundError:  # Python 3.13+ removed cgi.
    cgi = None


HOST = "127.0.0.1"
PORT = 8765
UPLOAD_DIR = OUT_DIR / "uploads"
STARTUP_LOG = OUT_DIR / "server_startup.log"
MAX_UPLOAD_BYTES = 80 * 1024 * 1024
DATA_CACHE: dict[tuple, dict] = {}
SUPABASE_CACHE: dict[str, object] = {"prepared": None, "loaded_at": 0.0, "source_label": ""}
SUPABASE_CACHE_LOCK = threading.Lock()
ACTIVE_CACHE: dict[tuple, dict[str, object]] = {}
ACTIVE_CACHE_LOCK = threading.Lock()
BACKGROUND_WARMUP_LOCK = threading.Lock()
BACKGROUND_WARMUP_STATE: dict[str, object] = {"running": False, "ready": False, "error": ""}
TAB_PAYLOAD_CACHE: dict[tuple, dict] = {}
TAB_PAYLOAD_CACHE_LOCK = threading.Lock()
ACTIVE_TAB_SNAPSHOT = OUT_DIR / "active_tab_snapshot.json"
SALES_TAB_SNAPSHOT = OUT_DIR / "sales_tab_snapshot.json"
SUPABASE_SOURCE_PATH = "supabase://basededadosEVO/public"
OPENAI_RESPONSES_URL = "https://api.openai.com/v1/responses"
DEFAULT_OPENAI_MODEL = "gpt-5.5"
PROJECT_DIR = Path(__file__).resolve().parent
STATIC_HTML_OUT = PROJECT_DIR / HTML_OUT.name
BUILDER_SOURCE = PROJECT_DIR / "build_biofisic_tabs_dashboard.py"
DARK_THEME_LOCK_ID = "biofisic-dark-theme-lock"
DARK_THEME_LOCK_CSS = """
<style id="biofisic-dark-theme-lock">
@media screen {
  :root {
    --bg: #050b0d !important;
    --panel: rgba(14, 27, 30, .90) !important;
    --panel-2: rgba(7, 18, 21, .88) !important;
    --green: #00f529 !important;
    --blue: #36a8ff !important;
    --red: #ff514d !important;
    --orange: #ffbd22 !important;
    --violet: #b16cff !important;
    --muted: #a9bac7 !important;
    --ink: #f6fbff !important;
    --line: rgba(0, 245, 41, .28) !important;
  }
  body {
    color: var(--ink) !important;
    background:
      linear-gradient(rgba(0, 245, 41, .045) 1px, transparent 1px),
      linear-gradient(90deg, rgba(0, 245, 41, .045) 1px, transparent 1px),
      radial-gradient(circle at 78% 8%, rgba(0, 245, 41, .18), transparent 30%),
      radial-gradient(circle at 18% 18%, rgba(54, 168, 255, .12), transparent 26%),
      linear-gradient(180deg, #071113 0%, #03080a 100%) !important;
    background-size: 84px 84px, 84px 84px, auto, auto, auto !important;
  }
  body::before {
    background: linear-gradient(115deg, rgba(0, 245, 41, .08), transparent 34%, rgba(54, 168, 255, .07)) !important;
    box-shadow: inset 0 0 120px rgba(0,0,0,.58) !important;
  }
  .topbar,
  .hero,
  .card,
  .panel,
  .composition-panel,
  .medal-board-panel,
  .brief-card,
  .benchmark-card,
  .cluster-unit-table-wrap {
    border-color: rgba(0, 245, 41, .22) !important;
    background: linear-gradient(145deg, rgba(18, 31, 34, .92), rgba(7, 17, 20, .90)) !important;
    color: var(--ink) !important;
  }
  .topbar {
    background: rgba(13, 22, 24, .90) !important;
  }
  h1 {
    color: var(--green) !important;
    font-style: italic !important;
    text-shadow: 0 0 22px rgba(0, 245, 41, .22) !important;
  }
  .tabs button {
    color: #d5e4ec !important;
    background: rgba(7, 18, 21, .86) !important;
    border-color: rgba(0, 245, 41, .28) !important;
  }
  .tabs button.active,
  .tabs button:hover,
  .file-button,
  .analyze-btn,
  .print-btn,
  .apply-filters-btn,
  .isaias-chat button {
    color: #021008 !important;
    background: linear-gradient(135deg, #00f529, #21d981) !important;
    border-color: rgba(0, 245, 41, .72) !important;
  }
  .filter-field input:not([type="checkbox"]),
  .filter-field select,
  .multi-select-toggle,
  .isaias-chat textarea,
  .multi-select-menu,
  .isaias-answer,
  .line-canvas,
  .timeline-chart,
  .access-day-card {
    color: var(--ink) !important;
    background: rgba(5, 15, 18, .92) !important;
    border-color: rgba(0, 245, 41, .24) !important;
  }
  .card strong,
  .panel h2,
  .brief-card h3,
  .benchmark-card h3,
  .isaias-chat h2,
  .cluster-value,
  .cluster-unit-name,
  .cluster-unit-cell strong,
  .cluster-unit-total,
  .bar-value,
  .multi-value,
  .legend-row strong {
    color: var(--ink) !important;
  }
  .source,
  .panel-subtitle,
  .card small,
  .brief-card p,
  .benchmark-card p,
  .isaias-chat p,
  .legend-row,
  .bar-label {
    color: var(--muted) !important;
  }
  .donut::after {
    background: #071113 !important;
  }
  .bar-track,
  .cluster-track,
  .cluster-unit-mini-track {
    background: rgba(166, 184, 196, .18) !important;
  }
}
</style>
""".strip()


def load_local_openai_env() -> None:
    """Load OpenAI credentials from local env files without overriding OS env vars."""
    env_paths = [
        PROJECT_DIR / ".env",
        PROJECT_DIR / "openai.env",
        OUT_DIR / "openai.env",
    ]
    allowed_keys = {"OPENAI_API_KEY", "OPENAI_MODEL"}
    for env_path in env_paths:
        if not env_path.exists():
            continue
        for raw_line in env_path.read_text(encoding="utf-8", errors="ignore").splitlines():
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key in allowed_keys and value and not os.environ.get(key):
                os.environ[key] = value


load_local_openai_env()

ISAIAS_SYSTEM_PROMPT = """
Voce e a isa IA, uma camada analitica critica para uma rede de academias BioFisic.
Responda em portugues do Brasil, com linguagem executiva, objetiva e acionavel.

Use como estilo profissional uma mescla de:
- Vinicius Ribeiro: executivo fitness com 18+ anos em operacao, produto, performance de unidades, abertura de academias, validacao de ponto, layout, midia, treinamento, padronizacao e resultado sustentavel.
- Junior Brandao: consultor estrategico de fitness no Brasil, com foco em expansao, operacoes, escalabilidade, franquias, estrutura comercial e leitura pratica de rede.

Regras:
- Use os numeros enviados no contexto do dashboard como fonte principal.
- Quando usar busca web, use apenas como benchmark de mercado, nao para substituir os dados da BioFisic.
- Seja critico: aponte risco, causa provavel, impacto e acao recomendada.
- Se o dado pedido nao estiver no contexto, diga que nao consta no dashboard atual e sugira qual CSV/grafico deveria alimentar essa resposta.
- Evite respostas genericas. Sempre que possivel, cite unidades, indicadores, percentuais, tickets, churn, frequencia, inadimplencia ou faturamento.
- Estruture respostas longas em: Diagnostico, Leitura executiva, Acao recomendada.
- Nao invente dados de outras redes. Se a busca nao trouxer referencia confiavel, diga que a comparacao externa ficou limitada.
""".strip()

TAB_LABELS = {
    "ativos": "Ativos",
    "vendas": "Vendas",
    "cancelamentos": "Cancelamentos",
    "financeiro": "Financeiro",
    "frequencia": "Frequencia",
    "isaias": "Analise",
}

HEADER_FILL = PatternFill("solid", fgColor="0B1C20")
SECTION_FILL = PatternFill("solid", fgColor="00F529")
SUBSECTION_FILL = PatternFill("solid", fgColor="D9EAF2")
WHITE_FONT = Font(color="FFFFFF", bold=True)
DARK_FONT = Font(color="031007", bold=True)
BOLD_FONT = Font(bold=True)
AUTO_PERCENT = "__AUTO_PERCENT__"


def startup_log(message: str) -> None:
    try:
        OUT_DIR.mkdir(parents=True, exist_ok=True)
        with STARTUP_LOG.open("a", encoding="utf-8") as handle:
            handle.write(f"{datetime.now():%Y-%m-%d %H:%M:%S} {message}\n")
    except Exception:
        pass


def source_signature(input_path: Path) -> tuple:
    resolved = input_path.resolve()
    if resolved.is_file():
        stat = resolved.stat()
        return (str(resolved), stat.st_mtime_ns, stat.st_size)
    files = []
    for item in sorted(resolved.glob("*")):
        if item.is_file() and item.suffix.lower() in {".csv", ".xlsx", ".zip"}:
            stat = item.stat()
            files.append((item.name.lower(), stat.st_mtime_ns, stat.st_size))
    return (str(resolved), tuple(files))


def get_processed_data(input_path: Path) -> dict:
    signature = source_signature(input_path)
    cached = DATA_CACHE.get(signature)
    if cached is not None:
        return cached
    prepared = load_processed_tables(input_path)
    DATA_CACHE.clear()
    DATA_CACHE[signature] = prepared
    return prepared


def supabase_cache_ttl_seconds() -> int:
    try:
        return max(30, int(os.environ.get("SUPABASE_CACHE_TTL_SECONDS", "300")))
    except ValueError:
        return 300


def get_supabase_processed_data(force: bool = False) -> tuple[dict, str]:
    """Read all dashboard tables from Supabase once, then reuse them across filters."""
    with SUPABASE_CACHE_LOCK:
        prepared = SUPABASE_CACHE.get("prepared")
        loaded_at = float(SUPABASE_CACHE.get("loaded_at") or 0.0)
        if prepared is not None and not force:
            source_label = str(SUPABASE_CACHE.get("source_label") or "Supabase")
            if time.time() - loaded_at >= supabase_cache_ttl_seconds():
                # Serve the last valid snapshot immediately and refresh it without
                # blocking a tab request for tens of seconds.
                start_background_warmup(force_refresh=True)
            return prepared, source_label

        seed_frames: dict[str, pd.DataFrame] = {}
        with ACTIVE_CACHE_LOCK:
            recent_active = max(
                ACTIVE_CACHE.values(),
                key=lambda item: float(item.get("loaded_at") or 0),
                default=None,
            )
            if recent_active:
                cached_frames = recent_active.get("frames") or {}
                for role in ("active", "active_history"):
                    if role in cached_frames:
                        seed_frames[role] = cached_frames[role]
        config = SupabaseConfig.from_environment(PROJECT_DIR)
        all_roles = {
            "active", "active_history", "billing", "sales", "sales_realtime", "cancellations", "non_renewed", "charges",
            "access_unit", "access_wellhub", "access_totalpass",
        }
        frames, validation = SupabaseTableLoader(config).fetch_roles(all_roles.difference(seed_frames))
        for role, frame in seed_frames.items():
            frames[role] = frame
            validation.append({
                "arquivo": role,
                "papel": role,
                "status": "cache ativo",
                "colunas": int(len(frame.columns)),
                "linhas": int(len(frame.index)),
            })
        prepared = prepare_table_frames(
            frames,
            source_path=SUPABASE_SOURCE_PATH,
            validation=validation,
        )
        if not prepared.get("valid"):
            missing = ", ".join(prepared.get("missingRoles", []))
            raise RuntimeError(f"As tabelas obrigatorias nao foram carregadas: {missing}")
        source_label = supabase_source_label()
        SUPABASE_CACHE.update(
            prepared=prepared,
            loaded_at=time.time(),
            source_label=source_label,
        )
        with TAB_PAYLOAD_CACHE_LOCK:
            TAB_PAYLOAD_CACHE.clear()
        return prepared, source_label


def active_period_window(filters: dict | None) -> tuple[str, str]:
    filters = filters or {}
    start = str(filters.get("periodStart") or "").strip()
    end = str(filters.get("periodEnd") or "").strip()
    if not start and not end:
        today = date.today()
        start = today.replace(day=1).isoformat()
        end = today.isoformat()
    return start, end


def get_supabase_active_data(filters: dict | None = None, force: bool = False) -> tuple[dict, str]:
    """Load the active dashboard first, using only the rows it needs."""
    start, end = active_period_window(filters)
    cache_key = (start, end)
    # Se o aquecimento das demais abas já terminou, a base completa também
    # contém tudo de que Ativos precisa. Reutilizá-la evita uma segunda rodada
    # de paginação quando o usuário volta para a primeira aba ou recarrega.
    with SUPABASE_CACHE_LOCK:
        full_prepared = SUPABASE_CACHE.get("prepared")
        full_loaded_at = float(SUPABASE_CACHE.get("loaded_at") or 0.0)
        full_source = str(SUPABASE_CACHE.get("source_label") or "Supabase")
        if (
            full_prepared is not None
            and not force
            and time.time() - full_loaded_at < supabase_cache_ttl_seconds()
        ):
            return full_prepared, full_source
    with ACTIVE_CACHE_LOCK:
        cached = ACTIVE_CACHE.get(cache_key)
        if cached and not force and time.time() - float(cached.get("loaded_at") or 0) < supabase_cache_ttl_seconds():
            return cached["prepared"], str(cached["source_label"])

        access_filters: list[tuple[str, str]] = []
        if start:
            access_filters.append(("date_event", f"gte.{start} 00:00:00"))
        if end:
            try:
                exclusive_end = (date.fromisoformat(end) + timedelta(days=1)).isoformat()
            except ValueError:
                exclusive_end = end
            access_filters.append(("date_event", f"lt.{exclusive_end} 00:00:00"))
        roles = {
            "active",
            "active_history",
            "charges",
            "access_unit",
            "access_wellhub",
            "access_totalpass",
        }
        role_filters = {
            "charges": [("status", "eq.a receber")],
            "access_unit": access_filters,
            "access_wellhub": access_filters,
            "access_totalpass": access_filters,
        }
        config = SupabaseConfig.from_environment(PROJECT_DIR)
        frames, validation = SupabaseTableLoader(config).fetch_roles(roles, role_filters)
        frames.setdefault("sales", pd.DataFrame())
        frames.setdefault("cancellations", pd.DataFrame())
        prepared = prepare_table_frames(
            frames,
            source_path=SUPABASE_SOURCE_PATH,
            validation=validation,
        )
        if not prepared.get("valid"):
            missing = ", ".join(prepared.get("missingRoles", []))
            raise RuntimeError(f"As tabelas da página Ativos não foram carregadas: {missing}")
        source_label = supabase_source_label()
        ACTIVE_CACHE[cache_key] = {
            "prepared": prepared,
            "frames": frames,
            "loaded_at": time.time(),
            "source_label": source_label,
        }
        return prepared, source_label


def _warm_full_supabase_cache(force_refresh: bool = False) -> None:
    try:
        get_supabase_processed_data(force=force_refresh)
        with BACKGROUND_WARMUP_LOCK:
            BACKGROUND_WARMUP_STATE.update(running=False, ready=True, error="")
        startup_log("cache completo do Supabase carregado em segundo plano")
    except Exception as exc:
        with BACKGROUND_WARMUP_LOCK:
            BACKGROUND_WARMUP_STATE.update(running=False, ready=False, error=str(exc))
        startup_log(f"falha no carregamento em segundo plano: {exc!r}")


def start_background_warmup(force_refresh: bool = False) -> None:
    with BACKGROUND_WARMUP_LOCK:
        if BACKGROUND_WARMUP_STATE.get("running") or (BACKGROUND_WARMUP_STATE.get("ready") and not force_refresh):
            return
        BACKGROUND_WARMUP_STATE.update(running=True, ready=False, error="")
    threading.Thread(
        target=_warm_full_supabase_cache,
        args=(force_refresh,),
        name="supabase-background-warmup",
        daemon=True,
    ).start()


def force_dark_theme_html(html: str) -> str:
    # O tema agora é controlado pelo usuário no próprio dashboard.
    return html


def tab_response(payload: dict, tab_key: str | None) -> dict:
    tabs = payload.get("tabs", {})
    selected = tab_key if tab_key in tabs else "ativos"
    return {
        "tabKey": selected,
        "tab": tabs.get(selected, {}),
        "filters": payload.get("filters", {}),
        "filterOptions": payload.get("filterOptions", {}),
        "sourceFile": payload.get("sourceFile", ""),
        "sourcePath": payload.get("sourcePath", ""),
        "medalBoard": payload.get("medalBoard", []),
        "validation": payload.get("validation", []),
    }


def save_tab_snapshot(response_payload: dict, target: Path, label: str) -> None:
    try:
        target.write_text(json.dumps(response_payload, ensure_ascii=False), encoding="utf-8")
    except OSError as exc:
        startup_log(f"nao foi possivel salvar snapshot de {label}: {exc!r}")


def tab_cache_key(tab_key: str, filters: dict) -> tuple:
    filter_keys = (
        "periodStart", "periodEnd", "unitFilter", "unitFilters",
        "ageFilter", "ageFilters", "genderFilter", "genderFilters",
    )
    return (
        tab_key,
        float(SUPABASE_CACHE.get("loaded_at") or 0.0),
        float(ANALYSIS_SNAPSHOT_FILE.stat().st_mtime if tab_key == "isaias" and ANALYSIS_SNAPSHOT_FILE.exists() else 0.0),
        *(str(filters.get(key) or "") for key in filter_keys),
    )


def read_response_text(response_payload: dict) -> str:
    text = response_payload.get("output_text")
    if isinstance(text, str) and text.strip():
        return text.strip()
    parts = []
    for item in response_payload.get("output", []) or []:
        for content in item.get("content", []) or []:
            if isinstance(content, dict):
                value = content.get("text")
                if isinstance(value, str) and value.strip():
                    parts.append(value.strip())
    return "\n\n".join(parts).strip()


def compact_json(value, max_chars: int = 28000) -> str:
    text = json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    if len(text) <= max_chars:
        return text
    return text[:max_chars] + "\n...[contexto reduzido por tamanho]"


ISAIAS_STOPWORDS = {
    "a", "ao", "aos", "as", "ate", "com", "como", "da", "das", "de", "do", "dos", "e", "em",
    "essa", "esse", "esta", "este", "eu", "me", "na", "nas", "no", "nos", "o", "os", "ou",
    "para", "por", "qual", "quais", "que", "se", "sobre", "tem", "ter", "um", "uma",
}

ISAIAS_INTENTS = {
    "financeiro": ("fatur", "receita", "caixa", "ticket", "receb", "pagamento", "mensalidade", "cobranca"),
    "vendas": ("venda", "contrato", "comercial", "checkout", "sv plus", "sangue verde", "plano"),
    "retencao": ("cancel", "churn", "retenc", "ltv", "risco", "abandono", "90", "120"),
    "frequencia": ("frequ", "acesso", "visita", "entrada", "horario", "wellhub", "totalpass", "agregador"),
    "perfil": ("perfil", "idade", "sexo", "faixa", "aluno", "cliente", "ativo", "inadimpl"),
    "unidades": ("unidade", "ranking", "melhor", "pior", "compar", "prioridade", "expans"),
}


def isaias_norm(value: object) -> str:
    text = str(value or "").lower()
    replacements = str.maketrans("áàâãäéèêëíìîïóòôõöúùûüç", "aaaaaeeeeiiiiooooouuuuc")
    return text.translate(replacements)


def isaias_tokens(value: object) -> set[str]:
    return {
        token
        for token in re.findall(r"[a-z0-9]{3,}", isaias_norm(value))
        if token not in ISAIAS_STOPWORDS
    }


def isaias_int(value: object) -> int:
    try:
        return int(float(value or 0))
    except (TypeError, ValueError):
        return 0


def isaias_br_int(value: object) -> str:
    return f"{isaias_int(value):,}".replace(",", ".")


def isaias_pct_detail(value: object) -> str:
    text = str(value or "0")
    return text if "%" in text else f"{text}%"


def isaias_number(value: object) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    raw_text = str(value)
    text = raw_text.replace(".", "").replace(",", ".") if "," in raw_text else raw_text
    match = re.search(r"-?\d+(?:[.,]\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def isaias_display_value(value: object) -> str:
    raw = str(value or "").strip()
    if "R$" in raw or "%" in raw:
        return raw
    number = isaias_number(value)
    if number is None:
        return raw
    if abs(number) >= 1000 and float(number).is_integer():
        return f"{int(number):,}".replace(",", ".")
    if float(number).is_integer():
        return f"{int(number)}"
    return f"{number:,.1f}".replace(",", "X").replace(".", ",").replace("X", ".")


def isaias_metric_text(item: dict) -> str:
    parts = [
        item.get("tab"),
        item.get("kind"),
        item.get("title"),
        item.get("label"),
        item.get("value"),
        item.get("display"),
        item.get("detail"),
        item.get("subtitle"),
    ]
    for bar in item.get("bars", []) or []:
        parts.extend([bar.get("label"), bar.get("display"), bar.get("value")])
    for cluster in item.get("clusters", []) or []:
        parts.extend([cluster.get("label"), cluster.get("value"), cluster.get("pct")])
    return " | ".join(str(part) for part in parts if part not in (None, ""))


def dashboard_semantic_items(dashboard_context: dict | None) -> list[dict]:
    dashboard_context = dashboard_context or {}
    tabs = dashboard_context.get("tabs", {}) or {}
    chat_context = dashboard_context.get("chatContext") or (
        tabs.get("isaias", {}).get("chatContext", {}) if isinstance(tabs, dict) else {}
    )
    items: list[dict] = []
    if chat_context:
        summary_pairs = [
            ("Alunos ativos", chat_context.get("active"), "base ativa atual"),
            ("Adimplentes", chat_context.get("adimplentes"), isaias_pct_detail(chat_context.get("adimplentesPct", 0))),
            ("Inadimplentes", chat_context.get("inadimplentes"), isaias_pct_detail(chat_context.get("inadimplentesPct", 0))),
            ("Contratos vendidos", chat_context.get("sales"), f"{chat_context.get('salesClients', 0)} clientes"),
            ("Cancelamentos", chat_context.get("cancellations"), f"{chat_context.get('cancelClients', 0)} clientes"),
            ("Acessos", chat_context.get("access"), f"{chat_context.get('accessClients', 0)} clientes"),
            ("Ticket vendido", chat_context.get("salesTicket"), "ticket medio de venda"),
            ("Ticket recebido", chat_context.get("receivedTicket"), "ticket medio de recebimento"),
            ("SV Plus", chat_context.get("svPlusPct"), "mix premium nas vendas"),
            ("Maior faturamento", chat_context.get("topRevenueUnit"), chat_context.get("topRevenueDisplay")),
            ("Melhor sucesso de venda", chat_context.get("topSalesSuccessUnit"), chat_context.get("topSalesSuccessDisplay")),
            ("Menor churn", chat_context.get("bestChurnUnit"), chat_context.get("bestChurnDisplay")),
            ("Cobranca critica", chat_context.get("paymentAlert"), chat_context.get("paymentAlertDisplay")),
        ]
        for label, value, detail in summary_pairs:
            if value not in (None, ""):
                items.append({
                    "tab": "Resumo executivo",
                    "kind": "indicador",
                    "title": label,
                    "label": label,
                    "value": value,
                    "detail": detail,
                })
    for row in dashboard_context.get("medalBoard", []) or []:
        items.append({
            "tab": "Ranking",
            "kind": "medalhas",
            "title": "Quadro de medalhas",
            "label": row.get("unit"),
            "value": f"ouro {row.get('gold', 0)}, prata {row.get('silver', 0)}, bronze {row.get('bronze', 0)}",
            "detail": f"total {row.get('total', 0)}",
        })
    for tab_key, tab in tabs.items():
        if not isinstance(tab, dict):
            continue
        for card_item in (tab.get("cards", []) or []):
            items.append({
                "tab": tab_key,
                "kind": "card",
                "title": card_item.get("label"),
                "label": card_item.get("label"),
                "value": card_item.get("value"),
                "detail": card_item.get("detail") or card_item.get("sub"),
            })
        for chart in (tab.get("charts", []) or []):
            title = chart.get("title") or "Grafico"
            subtitle = chart.get("subtitle") or ""
            for row in (chart.get("rows", []) or []):
                if not isinstance(row, dict):
                    continue
                items.append({
                    "tab": tab_key,
                    "kind": "grafico",
                    "title": title,
                    "subtitle": subtitle,
                    "label": row.get("label") or row.get("unit") or row.get("name"),
                    "value": row.get("display") or row.get("value") or row.get("total") or row.get("pct"),
                    "display": row.get("display") or row.get("medianDisplay") or row.get("meanDisplay"),
                    "detail": row.get("detail") or row.get("sub"),
                    "bars": row.get("bars", []),
                    "clusters": row.get("clusters", []),
                })
        composition = tab.get("composition", []) or []
        if isinstance(composition, dict):
            composition_rows = composition.get("rows", []) or []
        else:
            composition_rows = composition
        for row in composition_rows:
            if not isinstance(row, dict):
                continue
            items.append({
                "tab": tab_key,
                "kind": "composicao",
                "title": "Composicao",
                "label": row.get("label"),
                "value": row.get("display") or row.get("value"),
                "detail": row.get("detail"),
            })
    for item in items:
        item["_text"] = isaias_metric_text(item)
        item["_tokens"] = isaias_tokens(item["_text"])
    return [item for item in items if item.get("_text")]


def isaias_rank_items(question: str, items: list[dict]) -> list[dict]:
    query_tokens = isaias_tokens(question)
    query_norm = isaias_norm(question)
    want_low = any(term in query_norm for term in ("menor", "menos", "baixo", "baixa"))
    ranked = []
    for item in items:
        text_norm = isaias_norm(item.get("_text", ""))
        overlap = len(query_tokens & item.get("_tokens", set()))
        phrase_score = sum(2 for token in query_tokens if token and token in text_norm)
        tab_score = 2 if item.get("tab") and isaias_norm(item.get("tab")) in query_norm else 0
        score = overlap * 4 + phrase_score + tab_score
        if score:
            number = isaias_number(item.get("display") or item.get("value"))
            ranked.append((score, number if number is not None else -1.0, item))
    ranked.sort(key=lambda pair: (-pair[0], pair[1] if want_low else -pair[1]))
    return [item for _, _, item in ranked[:10]] if ranked else items[:8]


def isaias_detect_intent(question: str) -> str:
    q = isaias_norm(question)
    scores = {name: sum(1 for term in terms if term in q) for name, terms in ISAIAS_INTENTS.items()}
    best = max(scores.items(), key=lambda item: item[1])
    return best[0] if best[1] else "geral"


def isaias_interpretation(intent: str) -> str:
    messages = {
        "financeiro": "Leitura critica: se receita, recebimento e frequencia nao caminham juntos, o faturamento pode esconder fragilidade de caixa.",
        "vendas": "Leitura critica: volume comercial so e crescimento quando vira aluno ativo, frequente e pagante.",
        "retencao": "Leitura critica: cancelamento e sintoma tardio. O risco nasce antes, na combinacao de baixa visita, inadimplencia e pouco tempo ativo.",
        "frequencia": "Leitura critica: frequencia e a prova de valor percebido. Aluno que compra e nao entra precisa ser tratado como pre-churn.",
        "perfil": "Leitura critica: perfil so vira decisao quando conectado a comportamento: idade, sexo, canal e status financeiro devem orientar acao por unidade.",
        "unidades": "Leitura critica: a melhor unidade nao e necessariamente a que mais vende; e a que combina venda, recebimento, frequencia, baixa inadimplencia e menor churn.",
    }
    return messages.get(intent, "Leitura critica: o dashboard deve ser lido como sistema. Venda cria entrada, frequencia prova valor, cobranca confirma receita e cancelamento mostra onde a operacao chegou tarde.")


def isaias_recommendation(intent: str) -> str:
    messages = {
        "financeiro": "Proximo passo: separar receita confirmada de venda de entrada, acompanhar recuperacao de meses anteriores e criar rotina diaria para parcelas vencidas do mes.",
        "vendas": "Proximo passo: avaliar quais unidades vendem com maior sucesso depois da compra, replicar o playbook e auditar ofertas que geram inadimplencia.",
        "retencao": "Proximo passo: montar lista de risco com alunos sem visita, inadimplentes e recem-vendidos, atacando antes do pedido de cancelamento.",
        "frequencia": "Proximo passo: cruzar horario, unidade e canal para ajustar equipe, campanhas de ativacao e conversao de agregadores em relacionamento recorrente.",
        "perfil": "Proximo passo: transformar perfil em acao segmentada por faixa etaria, sexo, unidade, canal e comportamento de acesso.",
        "unidades": "Proximo passo: comparar as tres melhores e as tres piores na mesma metrica e buscar o processo operacional que explica a diferenca.",
    }
    return messages.get(intent, "Proximo passo: formule a decisao desejada com unidade, periodo e metrica. A isa IA local consegue cruzar os indicadores carregados sem custo de API.")


def local_isaias_answer(question: str, dashboard_context: dict | None = None, reason: str = "") -> str:
    dashboard_context = dashboard_context or {}
    chat_context = dashboard_context.get("chatContext") or {}
    if not chat_context:
        chat_context = (
            dashboard_context.get("tabs", {})
            .get("isaias", {})
            .get("chatContext", {})
        )
    question_text = (question or "").strip()
    active = chat_context.get("active", 0) or 0
    sales = chat_context.get("sales", 0) or 0
    cancellations = chat_context.get("cancellations", 0) or 0
    access = chat_context.get("access", 0) or 0
    received_ticket = chat_context.get("receivedTicket") or "sem leitura"
    sales_ticket = chat_context.get("salesTicket") or "sem leitura"
    top_revenue_unit = chat_context.get("topRevenueUnit") or "sem unidade"
    top_revenue_display = chat_context.get("topRevenueDisplay") or "sem leitura"

    context_line = (
        f"Base atual do dashboard: {int(active):,} ativos, {int(sales):,} contratos vendidos, "
        f"{int(cancellations):,} cancelamentos e {int(access):,} acessos."
    ).replace(",", ".")
    if not question_text:
        return "Digite uma pergunta para eu analisar os dados ou o contexto do mercado fitness."

    dashboard_terms = (
        "venda", "fatur", "receita", "ticket", "cancel", "churn", "reten",
        "inadimpl", "cobran", "frequ", "acesso", "unidade", "plano", "ltv",
        "wellhub", "totalpass", "aluno", "cliente", "biofisic", "mercado fitness",
    )
    is_dashboard_related = any(term in question_text.lower() for term in dashboard_terms)
    if not is_dashboard_related:
        prefix = (
            "Essa pergunta parece fugir do contexto direto do dashboard. "
            "Sem uma chave OPENAI_API_KEY ativa, eu respondo em modo local e não faço busca externa em tempo real."
        )
        if reason:
            prefix += f" Motivo técnico: {reason}"
        return (
            f"{prefix}\n\n"
            "Posso ainda ajudar de forma executiva: conecte a pergunta a uma decisão de operação, vendas, retenção, "
            "cobrança, expansão ou experiência do aluno. Com o ChatGPT configurado, a isa IA passa a responder "
            "perguntas gerais e comparar com referências externas.\n\n"
            f"{context_line}"
        )

    return (
        f"{context_line}\n\n"
        f"Leitura local: ticket vendido {sales_ticket}, ticket recebido {received_ticket}, "
        f"maior faturamento em {top_revenue_unit} ({top_revenue_display}). "
        "Para uma resposta mais aberta com benchmark de mercado e raciocínio fora do dashboard, configure a OPENAI_API_KEY."
    )


def local_isaias_answer(question: str, dashboard_context: dict | None = None, reason: str = "") -> str:
    dashboard_context = dashboard_context or {}
    chat_context = dashboard_context.get("chatContext") or {}
    if not chat_context:
        chat_context = (
            dashboard_context.get("tabs", {})
            .get("isaias", {})
            .get("chatContext", {})
        )
    question_text = (question or "").strip()
    active = chat_context.get("active", 0) or 0
    sales = chat_context.get("sales", 0) or 0
    cancellations = chat_context.get("cancellations", 0) or 0
    access = chat_context.get("access", 0) or 0
    received_ticket = chat_context.get("receivedTicket") or "sem leitura"
    sales_ticket = chat_context.get("salesTicket") or "sem leitura"
    top_revenue_unit = chat_context.get("topRevenueUnit") or "sem unidade"
    top_revenue_display = chat_context.get("topRevenueDisplay") or "sem leitura"

    context_line = (
        f"Base atual do dashboard: {int(active):,} ativos, {int(sales):,} contratos vendidos, "
        f"{int(cancellations):,} cancelamentos e {int(access):,} acessos."
    ).replace(",", ".")
    if not question_text:
        return "Digite uma pergunta para eu analisar os dados carregados do dashboard."

    items = dashboard_semantic_items(dashboard_context)
    matches = isaias_rank_items(question_text, items)
    intent = isaias_detect_intent(question_text)
    evidence_lines = []
    seen = set()
    for item in matches:
        label = str(item.get("label") or item.get("title") or "").strip()
        value = isaias_display_value(item.get("display") or item.get("value") or "")
        detail = str(item.get("detail") or item.get("subtitle") or "").strip()
        tab = str(item.get("tab") or "dashboard").strip()
        line_key = (tab, label, value, detail)
        if not label or line_key in seen:
            continue
        seen.add(line_key)
        suffix = f" - {detail}" if detail and detail != value else ""
        evidence_lines.append(f"- {tab}: {label}: {value}{suffix}")
        if len(evidence_lines) >= 6:
            break

    if not evidence_lines:
        evidence_lines = [
            f"- Resumo executivo: ativos {isaias_br_int(active)}, vendas {isaias_br_int(sales)}, cancelamentos {isaias_br_int(cancellations)} e acessos {isaias_br_int(access)}.",
            f"- Financeiro: ticket vendido {sales_ticket}; ticket recebido {received_ticket}; maior faturamento em {top_revenue_unit} ({top_revenue_display}).",
        ]

    answer = [
        "Modo local semantico gratuito: analisei a pergunta usando apenas os dados carregados no dashboard, sem API externa e sem custo de tokens.",
        "",
        context_line,
        "",
        "O que encontrei nos dados:",
        *evidence_lines,
        "",
        isaias_interpretation(intent),
        "",
        isaias_recommendation(intent),
    ]
    if reason:
        answer.append("")
        answer.append(f"Observacao tecnica: resposta local usada por seguranca ({reason}).")
    return "\n".join(answer)


def ask_openai_isaias(question: str, dashboard_context: dict, history: list | None = None, use_search: bool = True) -> dict:
    api_key = os.environ.get("OPENAI_API_KEY", "").strip()
    if not api_key:
        return {
            "answer": local_isaias_answer(question, dashboard_context, "OPENAI_API_KEY ausente"),
            "mode": "local",
            "warning": "OPENAI_API_KEY ausente. Resposta local ativada.",
        }

    model = os.environ.get("OPENAI_MODEL", DEFAULT_OPENAI_MODEL).strip() or DEFAULT_OPENAI_MODEL
    history_text = compact_json(history or [], 6000)
    context_text = compact_json(dashboard_context or {}, 30000)
    user_prompt = (
        f"Pergunta do usuario:\n{question.strip()}\n\n"
        f"Historico recente do chat:\n{history_text}\n\n"
        f"Contexto numerico do dashboard:\n{context_text}\n\n"
        "Responda como consultor executivo de mercado fitness. "
        "Use benchmark externo apenas quando a busca estiver disponivel e for relevante."
    )
    payload = {
        "model": model,
        "instructions": ISAIAS_SYSTEM_PROMPT,
        "input": user_prompt,
        "max_output_tokens": 1400,
    }
    if use_search:
        payload["tools"] = [{"type": "web_search_preview"}]

    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = urlrequest.Request(
        OPENAI_RESPONSES_URL,
        data=data,
        method="POST",
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
    )
    try:
        with urlrequest.urlopen(req, timeout=60) as response:
            response_payload = json.loads(response.read().decode("utf-8"))
    except urlerror.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        if use_search and exc.code in {400, 404, 422}:
            return ask_openai_isaias(question, dashboard_context, history=history, use_search=False)
        raise ValueError(f"OpenAI API retornou {exc.code}: {detail[:500]}") from exc
    except urlerror.URLError as exc:
        raise ValueError(f"Nao foi possivel conectar a OpenAI: {exc.reason}") from exc

    answer = read_response_text(response_payload)
    if not answer:
        raise ValueError("A OpenAI respondeu sem texto utilizavel.")
    return {
        "answer": answer,
        "mode": "chatgpt_search" if use_search else "chatgpt",
        "model": model,
    }


def safe_sheet_name(name: str, used: set[str]) -> str:
    cleaned = re.sub(r"[:\\/?*\[\]]", " ", str(name)).strip() or "Unidade"
    cleaned = re.sub(r"\s+", " ", cleaned)[:31]
    base = cleaned
    counter = 2
    while cleaned in used:
        suffix = f" {counter}"
        cleaned = f"{base[:31 - len(suffix)]}{suffix}"
        counter += 1
    used.add(cleaned)
    return cleaned


def cell_value(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Sim" if value else "Nao"
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def normalize_text(value) -> str:
    text = str(value or "").strip().lower()
    replacements = {
        "á": "a", "à": "a", "ã": "a", "â": "a",
        "é": "e", "ê": "e",
        "í": "i",
        "ó": "o", "ô": "o", "õ": "o",
        "ú": "u",
        "ç": "c",
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    return text


def is_percent_header(value) -> bool:
    text = normalize_text(value)
    return text.startswith("%") or "percentual" in text


def export_percent_value(value):
    if value in (None, ""):
        return ""
    if isinstance(value, (int, float)):
        return f"={value}/100" if abs(value) > 1 else value
    return value


def write_row(ws, row_idx: int, values: list, fill=None, font=None) -> int:
    max_col = 5 if fill else min(5, len(values))
    for col_idx in range(1, max_col + 1):
        value = values[col_idx - 1] if col_idx <= len(values) else None
        cell = ws.cell(row=row_idx, column=col_idx, value=cell_value(value))
        cell.alignment = Alignment(
            horizontal="center" if col_idx > 1 else None,
            vertical="top",
            wrap_text=True,
        )
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
    return row_idx + 1


def write_table(ws, row_idx: int, title: str, headers: list[str], rows: list[list], gap: bool = True) -> int:
    if gap:
        row_idx += 1
    row_idx = write_row(ws, row_idx, [title], fill=SUBSECTION_FILL, font=BOLD_FONT)
    visible_headers = headers[:5]
    percent_cols = [idx + 1 for idx, header in enumerate(visible_headers) if is_percent_header(header)]
    row_idx = write_row(ws, row_idx, visible_headers, fill=HEADER_FILL, font=WHITE_FONT)
    if rows:
        for row in rows:
            current_row = row_idx
            visible_row = list(row[:5])
            auto_percent_seen = 0
            for col_idx in percent_cols:
                if col_idx > len(visible_row):
                    continue
                value = visible_row[col_idx - 1]
                if value == AUTO_PERCENT:
                    source_col = 2 + auto_percent_seen
                    total_last_col = 1 + len(percent_cols)
                    visible_row[col_idx - 1] = f"={get_column_letter(source_col)}{current_row}/SUM($B{current_row}:${get_column_letter(total_last_col)}{current_row})"
                    auto_percent_seen += 1
                else:
                    visible_row[col_idx - 1] = export_percent_value(value)
            row_idx = write_row(ws, row_idx, visible_row)
            for col_idx in percent_cols:
                ws.cell(row=current_row, column=col_idx).number_format = "0.0%"
    else:
        row_idx = write_row(ws, row_idx, ["Sem dados suficientes"])
    return row_idx


def card_rows(cards: list[dict]) -> list[list]:
    rows = []
    for card in cards or []:
        detail = card.get("sub") or card.get("meta") or ""
        rows.append([
            card.get("label", ""),
            card.get("value", ""),
            detail,
            card.get("metric", ""),
            card.get("status", ""),
        ])
    return rows


def chart_table(chart: dict) -> tuple[list[str], list[list]]:
    rows = chart.get("rows") or []
    chart_type = chart.get("type", "")
    if chart_type == "dualBar":
        primary = chart.get("primaryLabel") or "Mediana"
        secondary = chart.get("secondaryLabel") or "Media"
        return (
            ["Item", primary, secondary, f"{primary} exibido", f"{secondary} exibido", "Medalha"],
            [[
                row.get("label", ""),
                row.get("medianValue", ""),
                row.get("meanValue", ""),
                row.get("medianDisplay", ""),
                row.get("meanDisplay", ""),
                row.get("medal", ""),
            ] for row in rows],
        )
    if chart_type == "multiBar":
        labels = []
        for row in rows:
            for bar in row.get("bars") or []:
                label = bar.get("label", "Valor")
                if label not in labels:
                    labels.append(label)
        table_rows = []
        for row in rows:
            by_label = {bar.get("label", "Valor"): bar for bar in row.get("bars") or []}
            table_rows.append(
                [row.get("label", ""), row.get("medal", "")]
                + [by_label.get(label, {}).get("display", by_label.get(label, {}).get("value", "")) for label in labels]
            )
        return ["Item", "Medalha"] + labels, table_rows
    if chart_type in {"columnBar", "stackedColumn"} and rows and isinstance(rows[0].get("segments"), list):
        labels = []
        for row in rows:
            for segment in row.get("segments") or []:
                label = segment.get("label", "Valor")
                if label not in labels:
                    labels.append(label)
        table_rows = []
        for row in rows:
            by_label = {segment.get("label", "Valor"): segment for segment in row.get("segments") or []}
            values = [by_label.get(label, {}).get("value", "") for label in labels]
            if len(labels) == 2:
                table_rows.append([row.get("label", "")] + values + [AUTO_PERCENT, AUTO_PERCENT])
            else:
                table_rows.append([row.get("label", "")] + values)
        if len(labels) == 2:
            return ["Item"] + labels + [f"% {labels[0]}", f"% {labels[1]}"], table_rows
        return ["Item"] + labels, table_rows
    if chart_type == "lineChart":
        series = chart.get("series") or []
        return (
            ["Data"] + [item.get("label", item.get("key", "")) for item in series],
            [[row.get("label", "")] + [row.get(item.get("key", ""), "") for item in series] for row in rows],
        )
    if chart_type == "clusterPanel":
        return (
            ["Cluster", "Faixa", "Quantidade", "Percentual"],
            [[row.get("label", ""), row.get("range", ""), row.get("value", ""), row.get("pct", "")] for row in rows],
        )
    if chart_type == "clusterUnitTable":
        clusters = chart.get("clusters") or []
        headers = ["Unidade"] + [f"{item.get('label', '')} ({item.get('range', '')})" for item in clusters] + ["Total"]
        table_rows = []
        for row in rows:
            cluster_values = [
                f"{cluster.get('value', 0)} | {cluster.get('pct', 0):.1f}%"
                for cluster in row.get("clusters") or []
            ]
            table_rows.append([row.get("unit", "")] + cluster_values + [row.get("total", "")])
        return headers, table_rows
    include_medal = any(row.get("medal") for row in rows)
    headers = ["Item", "Valor", "Percentual", "Exibicao"] + (["Medalha"] if include_medal else [])
    table_rows = []
    for row in rows:
        values = [row.get("label", ""), row.get("value", ""), row.get("pct", ""), row.get("display", "")]
        if include_medal:
            values.append(row.get("medal", ""))
        table_rows.append(values)
    return headers, table_rows


def chart_row_unit_label(row: dict) -> str:
    return str(row.get("unit") or row.get("label") or "").strip()


def scoped_chart_for_unit(chart: dict, unit: str) -> dict:
    rows = chart.get("rows") or []
    if not rows:
        return chart
    unit_labels = set(UNIT_ORDER)
    unit_row_count = sum(1 for row in rows if chart_row_unit_label(row) in unit_labels)
    class_name = normalize_text(chart.get("className", ""))
    title = normalize_text(chart.get("title", ""))
    should_scope = unit_row_count > 1 or "por unidade" in title or "unit" in class_name
    if not should_scope:
        return chart
    filtered_rows = [row for row in rows if chart_row_unit_label(row) == unit]
    if not filtered_rows:
        return chart
    scoped = dict(chart)
    scoped["rows"] = filtered_rows
    return scoped


def write_dashboard_payload_sheet(ws, unit: str, payload: dict) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws["A1"] = f"Dashboard BioFisic - {unit}"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(vertical="center")
    row_idx = 2

    tabs = payload.get("tabs", {})
    first_tab = True
    for tab_key, tab_label in TAB_LABELS.items():
        if tab_key == "isaias":
            continue
        tab = tabs.get(tab_key) or {}
        if not first_tab:
            row_idx += 1
        first_tab = False
        row_idx = write_row(ws, row_idx, [tab_label], fill=SECTION_FILL, font=DARK_FONT)
        first_table = True
        cards = list(tab.get("cards") or []) + list(tab.get("aggregatorCards") or [])
        if cards:
            row_idx = write_table(ws, row_idx, f"{tab_label} - Cards", ["Indicador", "Valor", "Detalhe", "Percentual", "Status"], card_rows(cards), gap=False)
            first_table = False
        if tab.get("composition"):
            headers, rows = chart_table({"type": "donut", **tab["composition"]})
            row_idx = write_table(ws, row_idx, f"{tab_label} - {tab['composition'].get('title', 'Composicao')}", headers, rows, gap=not first_table)
            first_table = False
        for chart in tab.get("charts") or []:
            chart = scoped_chart_for_unit(chart, unit)
            headers, rows = chart_table(chart)
            title = chart.get("title", "Grafico")
            subtitle = chart.get("subtitle", "")
            row_idx = write_table(ws, row_idx, f"{tab_label} - {title}", headers, rows, gap=not first_table)
            first_table = False
        if tab_key == "isaias":
            for collection, title in [("briefing", "Briefing"), ("signals", "Sinais"), ("benchmarks", "Benchmarks")]:
                items = tab.get(collection) or []
                if not items:
                    continue
                headers = sorted({key for item in items for key in item.keys()})
                row_idx = write_table(ws, row_idx, f"{tab_label} - {title}", headers, [[item.get(header, "") for header in headers] for item in items], gap=not first_table)
                first_table = False

    widths = {"A": 42.109375, "B": 25, "C": 44.5546875, "D": 21, "E": 23.77734375, "F": 28}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"


def build_unit_xlsx(
    input_path: Path | None,
    source_label: str,
    filters: dict,
    prepared_data: dict | None = None,
) -> bytes:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    used: set[str] = set()
    base_filters = dict(filters or {})
    base_filters.pop("unitFilter", None)
    base_filters.pop("unitFilters", None)
    if prepared_data is not None:
        prepared = prepared_data
    elif input_path is not None:
        prepared = get_processed_data(input_path)
    else:
        raise ValueError("Nenhuma fonte de dados foi informada para a exportacao.")
    for unit in UNIT_ORDER:
        unit_filters = {**base_filters, "unitFilter": unit}
        payload = build_payload(input_path, source_label=source_label, filters=unit_filters, prepared_data=prepared)
        sheet_name = safe_sheet_name(unit, used)
        ws = workbook.create_sheet(sheet_name)
        write_dashboard_payload_sheet(ws, unit, payload)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


@dataclass
class UploadItem:
    filename: str
    value: str
    file: io.BytesIO


class DashboardHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(OUT_DIR), **kwargs)

    def log_message(self, format, *args):  # noqa: A002
        print(f"[dashboard] {self.address_string()} - {format % args}")

    def end_headers(self):  # noqa: N802
        self.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
        self.send_header("Pragma", "no-cache")
        self.send_header("Expires", "0")
        super().end_headers()

    def do_GET(self):  # noqa: N802
        parsed = urlparse(self.path)
        if parsed.path == "/api/health":
            loaded_at = float(SUPABASE_CACHE.get("loaded_at") or 0.0)
            self._send_json({
                "status": "ok",
                "source": SUPABASE_SOURCE_PATH,
                "cacheLoaded": SUPABASE_CACHE.get("prepared") is not None,
                "cacheAgeSeconds": round(max(0.0, time.time() - loaded_at), 1) if loaded_at else None,
                "activeCacheLoaded": bool(ACTIVE_CACHE),
                "backgroundLoading": bool(BACKGROUND_WARMUP_STATE.get("running")),
                "backgroundReady": bool(BACKGROUND_WARMUP_STATE.get("ready")),
                "backgroundError": str(BACKGROUND_WARMUP_STATE.get("error") or ""),
            })
            return
        if parsed.path == "/":
            self.send_response(302)
            self.send_header("Location", "/dashboard_vendas_mar_abr_mai_2026.html")
            self.end_headers()
            return
        super().do_GET()

    def do_POST(self):  # noqa: N802
        parsed = urlparse(self.path)
        if parsed.path not in {"/api/supabase/sync", "/api/render-tab", "/api/export-xlsx", "/api/ask-isaias"}:
            self._send_json({"error": "Endpoint não encontrado."}, status=404)
            return

        try:
            length = int(self.headers.get("Content-Length", "0"))
        except ValueError:
            length = 0
        if length > MAX_UPLOAD_BYTES:
            self._send_json({"error": "Requisição muito grande."}, status=413)
            return

        if parsed.path == "/api/ask-isaias":
            try:
                request_payload = self._receive_json(length)
                question = str(request_payload.get("question") or "").strip()
                if not question:
                    self._send_json({"error": "Digite uma pergunta para a isa IA."}, status=400)
                    return
                self._send_json({
                    "answer": local_isaias_answer(question, request_payload.get("dashboard") or {}),
                    "mode": "local_semantic",
                })
            except Exception as exc:
                fallback_answer = local_isaias_answer(
                    question if "question" in locals() else "",
                    request_payload.get("dashboard") if "request_payload" in locals() else {},
                    str(exc),
                )
                self._send_json({
                    "answer": fallback_answer,
                    "mode": "local_semantic",
                    "warning": f"Resposta local ativada com contexto reduzido: {exc}",
                })
            return

        try:
            filters, force = self._receive_filters(length)
            requested_tab = filters.get("activeTab") if parsed.path == "/api/render-tab" else None
            if requested_tab == "ativos":
                if force:
                    with SUPABASE_CACHE_LOCK:
                        SUPABASE_CACHE.update(prepared=None, loaded_at=0.0, source_label="")
                    with BACKGROUND_WARMUP_LOCK:
                        BACKGROUND_WARMUP_STATE.update(ready=False, error="")
                prepared, source_label = get_supabase_active_data(filters, force=force)
            else:
                prepared, source_label = get_supabase_processed_data(
                    force=force if parsed.path == "/api/supabase/sync" else False
                )
            if parsed.path == "/api/export-xlsx":
                workbook_bytes = build_unit_xlsx(
                    None,
                    source_label,
                    filters,
                    prepared_data=prepared,
                )
                filename = f"dashboard_biofisic_por_unidade_{datetime.now():%Y%m%d_%H%M}.xlsx"
                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
                self.send_header("Content-Length", str(len(workbook_bytes)))
                self.end_headers()
                self.wfile.write(workbook_bytes)
                return
            if parsed.path == "/api/render-tab" and requested_tab != "ativos":
                cache_key = tab_cache_key(str(requested_tab), filters)
                with TAB_PAYLOAD_CACHE_LOCK:
                    response_payload = TAB_PAYLOAD_CACHE.get(cache_key)
                    if response_payload is None:
                        payload = build_payload(
                            None,
                            source_label=source_label,
                            filters=filters,
                            prepared_data=prepared,
                            only_tab=requested_tab,
                        )
                        response_payload = tab_response(payload, requested_tab)
                        TAB_PAYLOAD_CACHE[cache_key] = response_payload
                if requested_tab == "vendas":
                    save_tab_snapshot(response_payload, SALES_TAB_SNAPSHOT, "Vendas")
                self._send_json(response_payload)
                return

            payload = build_payload(
                None,
                source_label=source_label,
                filters=filters,
                prepared_data=prepared,
                only_tab=requested_tab if parsed.path == "/api/render-tab" else None,
            )
            if parsed.path == "/api/render-tab":
                response_payload = tab_response(payload, filters.get("activeTab"))
                if requested_tab == "ativos":
                    save_tab_snapshot(response_payload, ACTIVE_TAB_SNAPSHOT, "Ativos")
                self._send_json(response_payload)
                if requested_tab == "ativos":
                    start_background_warmup()
                return
            html = force_dark_theme_html(build_html(payload))
            HTML_OUT.write_text(html, encoding="utf-8")
            STATIC_HTML_OUT.write_text(html, encoding="utf-8")
            body = html.encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
        except Exception as exc:  # Keep errors user-readable in the dashboard.
            self._send_json({"error": f"Não consegui sincronizar o Supabase: {exc}"}, status=400)

    def _receive_filters(self, length: int) -> tuple[dict, bool]:
        content_type = self.headers.get("Content-Type", "")
        if length <= 0:
            form = {}
        elif "application/json" in content_type:
            form = self._receive_json(length)
        elif "multipart/form-data" in content_type:
            if cgi is not None:
                form = cgi.FieldStorage(
                    fp=self.rfile,
                    headers=self.headers,
                    environ={
                        "REQUEST_METHOD": "POST",
                        "CONTENT_TYPE": content_type,
                        "CONTENT_LENGTH": self.headers.get("Content-Length", "0"),
                    },
                )
            else:
                form = self._parse_multipart_form(content_type)
        else:
            raise ValueError("Envie os filtros em JSON ou multipart/form-data.")

        def value(key: str) -> str:
            if isinstance(form, dict) and key in form and not hasattr(form[key], "value"):
                item = form[key]
                if isinstance(item, list):
                    item = item[0] if item else ""
                return str(item or "").strip()
            return self._form_value(form, key)

        filters = {
            "periodStart": value("periodStart"),
            "periodEnd": value("periodEnd"),
            "unitFilter": value("unitFilter"),
            "unitFilters": value("unitFilters"),
            "ageFilter": value("ageFilter"),
            "ageFilters": value("ageFilters"),
            "genderFilter": value("genderFilter"),
            "genderFilters": value("genderFilters"),
            "activeTab": value("activeTab"),
        }
        force = value("force").lower() in {"1", "true", "yes", "sim"}
        return filters, force

    def _receive_json(self, length: int) -> dict:
        content_type = self.headers.get("Content-Type", "")
        if "application/json" not in content_type:
            raise ValueError("envie a pergunta em JSON.")
        body = self.rfile.read(length)
        if not body:
            return {}
        return json.loads(body.decode("utf-8", errors="replace"))

    def _receive_upload(self) -> tuple[Path, str, dict]:
        content_type = self.headers.get("Content-Type", "")
        if "multipart/form-data" not in content_type:
            raise ValueError("envie o pacote pelo seletor de arquivo do dashboard.")

        if cgi is not None:
            form = cgi.FieldStorage(
                fp=self.rfile,
                headers=self.headers,
                environ={
                    "REQUEST_METHOD": "POST",
                    "CONTENT_TYPE": content_type,
                    "CONTENT_LENGTH": self.headers.get("Content-Length", "0"),
                },
            )
        else:
            form = self._parse_multipart_form(content_type)
        filters = {
            "periodStart": self._form_value(form, "periodStart"),
            "periodEnd": self._form_value(form, "periodEnd"),
            "unitFilter": self._form_value(form, "unitFilter"),
            "unitFilters": self._form_value(form, "unitFilters"),
            "ageFilter": self._form_value(form, "ageFilter"),
            "ageFilters": self._form_value(form, "ageFilters"),
            "genderFilter": self._form_value(form, "genderFilter"),
            "genderFilters": self._form_value(form, "genderFilters"),
            "activeTab": self._form_value(form, "activeTab"),
        }
        if "workbook" not in form and "files" in form:
            input_path, source_label = self._save_csv_bundle(form["files"])
            return input_path, source_label, filters
        if "workbook" not in form:
            source_path = Path(self._form_value(form, "sourcePath") or DEFAULT_CSV_DIR)
            source_label = self._form_value(form, "sourceLabel")
            if not source_path.exists():
                source_path = DEFAULT_CSV_DIR
            if not source_label.startswith("Último upload:"):
                source_label = source_upload_label(source_path)
            return source_path, source_label, filters
        input_path, source_label = self._save_workbook(form["workbook"])
        return input_path, source_label, filters

    def _parse_multipart_form(self, content_type: str) -> dict[str, list[UploadItem]]:
        length = int(self.headers.get("Content-Length", "0"))
        body = self.rfile.read(length)
        raw = (
            f"Content-Type: {content_type}\r\n"
            "MIME-Version: 1.0\r\n\r\n"
        ).encode("utf-8") + body
        message = BytesParser(policy=email_policy).parsebytes(raw)
        form: dict[str, list[UploadItem]] = {}
        for part in message.iter_parts():
            disposition = part.get("Content-Disposition", "")
            if "form-data" not in disposition:
                continue
            name = part.get_param("name", header="content-disposition")
            if not name:
                continue
            filename = part.get_filename() or ""
            payload = part.get_payload(decode=True) or b""
            value = ""
            if not filename:
                charset = part.get_content_charset() or "utf-8"
                value = payload.decode(charset, errors="replace").strip()
            form.setdefault(str(name), []).append(UploadItem(filename=filename, value=value, file=io.BytesIO(payload)))
        return form

    def _form_value(self, form, key: str) -> str:
        if key not in form:
            return ""
        item = form[key]
        if isinstance(item, list):
            item = item[0]
        return str(getattr(item, "value", "") or "").strip()

    def _save_workbook(self, item) -> tuple[Path, str]:
        if isinstance(item, list):
            item = item[0]
        if not item.filename:
            raise ValueError("selecione um arquivo .xlsx.")

        original_name = Path(item.filename).name
        if not original_name.lower().endswith(".xlsx"):
            raise ValueError("o arquivo precisa estar no formato .xlsx.")

        UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        target = UPLOAD_DIR / original_name
        counter = 1
        while target.exists():
            target = UPLOAD_DIR / f"{Path(original_name).stem}_{counter}.xlsx"
            counter += 1

        with target.open("wb") as out:
            shutil.copyfileobj(item.file, out)
        return target, source_upload_label(target)

    def _save_csv_bundle(self, items) -> tuple[Path, str]:
        if not isinstance(items, list):
            items = [items]
        upload_items = [
            item
            for item in items
            if item.filename and Path(item.filename).suffix.lower() in {".csv", ".zip"}
        ]
        if not upload_items:
            raise ValueError("selecione os arquivos .csv ou um pacote .zip da analise.")

        UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        counter = 1
        target_dir = UPLOAD_DIR / "csv_bundle"
        while target_dir.exists():
            counter += 1
            target_dir = UPLOAD_DIR / f"csv_bundle_{counter}"
        target_dir.mkdir(parents=True, exist_ok=True)

        names = []
        for item in upload_items:
            original_name = Path(item.filename).name
            suffix = Path(original_name).suffix.lower()
            target = target_dir / original_name
            name_counter = 1
            while target.exists():
                target = target_dir / f"{Path(original_name).stem}_{name_counter}{suffix}"
                name_counter += 1
            with target.open("wb") as out:
                shutil.copyfileobj(item.file, out)
            names.append(original_name)
            if suffix == ".zip":
                self._extract_csv_zip(target, target_dir)
        return target_dir, source_upload_label(target_dir)

    def _extract_csv_zip(self, zip_path: Path, target_dir: Path) -> None:
        with zipfile.ZipFile(zip_path) as archive:
            for member in archive.infolist():
                member_name = Path(member.filename).name
                if not member_name or not member_name.lower().endswith(".csv"):
                    continue
                target = target_dir / member_name
                counter = 1
                while target.exists():
                    target = target_dir / f"{Path(member_name).stem}_{counter}.csv"
                    counter += 1
                with archive.open(member) as src, target.open("wb") as out:
                    shutil.copyfileobj(src, out)

    def _send_json(self, payload: dict, status: int = 200) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    startup_log("main iniciado")
    try:
        startup_log("gerando shell leve para carregamento progressivo")
        shell_payload = build_blank_payload(source_label="Conectando ao Supabase...")
        shell_payload["sourcePath"] = SUPABASE_SOURCE_PATH
        shell_payload["progressive"] = True
        for tab in shell_payload.get("tabs", {}).values():
            tab["loading"] = True
        builder_updated_at = BUILDER_SOURCE.stat().st_mtime if BUILDER_SOURCE.exists() else 0.0
        if ACTIVE_TAB_SNAPSHOT.exists() and ACTIVE_TAB_SNAPSHOT.stat().st_mtime >= builder_updated_at:
            try:
                active_snapshot = json.loads(ACTIVE_TAB_SNAPSHOT.read_text(encoding="utf-8"))
                if active_snapshot.get("tabKey") == "ativos" and active_snapshot.get("tab"):
                    shell_payload["tabs"]["ativos"] = active_snapshot["tab"]
                    shell_payload["sourceFile"] = active_snapshot.get("sourceFile") or shell_payload["sourceFile"]
                    shell_payload["sourcePath"] = active_snapshot.get("sourcePath") or shell_payload["sourcePath"]
                    shell_payload["validation"] = active_snapshot.get("validation") or []
                    shell_payload["activeSnapshot"] = True
                    startup_log("snapshot de Ativos incorporado ao shell")
            except (OSError, ValueError, TypeError) as exc:
                startup_log(f"snapshot de Ativos ignorado: {exc!r}")
        if SALES_TAB_SNAPSHOT.exists() and SALES_TAB_SNAPSHOT.stat().st_mtime >= builder_updated_at:
            try:
                sales_snapshot = json.loads(SALES_TAB_SNAPSHOT.read_text(encoding="utf-8"))
                if sales_snapshot.get("tabKey") == "vendas" and sales_snapshot.get("tab"):
                    shell_payload["tabs"]["vendas"] = sales_snapshot["tab"]
                    shell_payload["salesSnapshot"] = True
                    startup_log("snapshot de Vendas incorporado ao shell")
            except (OSError, ValueError, TypeError) as exc:
                startup_log(f"snapshot de Vendas ignorado: {exc!r}")
        initial_html = force_dark_theme_html(build_html(shell_payload))
        HTML_OUT.write_text(initial_html, encoding="utf-8")
        if not STATIC_HTML_OUT.exists():
            STATIC_HTML_OUT.write_text(initial_html, encoding="utf-8")
        startup_log("shell progressivo pronto; os dados serao carregados pelo navegador")
        startup_log(f"criando servidor {HOST}:{PORT}")
        server = ThreadingHTTPServer((HOST, PORT), DashboardHandler)
        startup_log("servidor pronto; entrando em serve_forever")
        print(f"Dashboard em http://{HOST}:{PORT}/dashboard_vendas_mar_abr_mai_2026.html")
        server.serve_forever()
    except Exception as exc:
        startup_log(f"erro no servidor: {exc!r}")
        raise


if __name__ == "__main__":
    main()
